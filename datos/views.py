from django.shortcuts import render
from .models import Datos
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side

def home(request):
    queryset = Datos.objects.all()
    return render(request,'index.html',{'queryset':queryset})

class ReportePersonalizadoExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        campo = int(request.GET.get('campo'))
        query = Datos.objects.filter(edad = campo)
        wb = Workbook()
        bandera = True
        cont = 1
        controlador = 4
        for q in query:
            if bandera:
                ws = wb.active
                ws.title = 'Hoja'+str(cont)
                bandera = False
            else:
                ws = wb.create_sheet('Hoja'+str(cont))
            #Crear el título en la hoja
            ws['B1'].alignment = Alignment(horizontal = "center",vertical = "center")
            ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") ) 
            ws['B1'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
            ws['B1'].font = Font(name = 'Calibri', size = 12, bold = True)
            ws['B1'] = 'REPORTE PERSONALIZADO EN EXCEL CON DJANGO'

            #Cambiar caracteristicas de las celdas
            ws.merge_cells('B1:E1')

            ws.row_dimensions[1].height = 25

            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20

            #Crear la cabecera
            ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws['B3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
            ws['B3'].font = Font(name = 'Calibro', size = 10, bold = True)
            ws['B3'] = 'Nombres'

            ws['C3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['C3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws['C3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
            ws['C3'].font = Font(name = 'Calibro', size = 10, bold = True)
            ws['C3'] = 'Apellidos'

            ws['D3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['D3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws['D3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
            ws['D3'].font = Font(name = 'Calibro', size = 10, bold = True)
            ws['D3'] = 'Dirección'

            ws['E3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['E3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws['E3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
            ws['E3'].font = Font(name = 'Calibro', size = 10, bold = True)
            ws['E3'] = 'Edad'


            #Pintamos los datos en el reporte
            ws.cell(row = controlador, column = 2).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = controlador, column = 2).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 2).value = q.nombre

            ws.cell(row = controlador, column = 3).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = controlador, column = 3).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 3).value = q.apellidos

            ws.cell(row = controlador, column = 4).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = controlador, column = 4).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 4).value = q.direccion


            ws.cell(row = controlador, column = 5).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = controlador, column = 5).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 5).value = q.edad
            
            cont += 1

        #Establecer el nombre de mi archivo
        nombre_archivo = "ReportePersonalizadoExcel.xlsx"
        #Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response